import pandas as pd
import os
from openpyxl import load_workbook
# 设置文件夹路径
folder_path = r'C:\Users\hjj\Desktop\2024年新闻稿件'  # 替换为你的文件夹路径
output_file = r'C:\Users\hjj\Desktop\新建文件夹\merged_data.xlsx'  # 输出文件名
image_folder = r'C:\Users\hjj\Desktop\新建文件夹\extracted_images'  # 保存图片的文件夹

# 创建保存图片的文件夹
if not os.path.exists(image_folder):
    os.makedirs(image_folder)

# 创建一个空的DataFrame，用于存储所有数据
all_data = pd.DataFrame(columns=['序号', '发布媒体', '媒体类型', '稿件名称', '相关链接'])

# 遍历文件夹中的所有文件
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx') or file_name.endswith('.xls'):  # 确保只处理Excel文件
        file_path = os.path.join(folder_path, file_name)
        try:
            # 加载工作簿
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active  # 获取活动工作表

            # 获取列名所在的行（假设第二行是列名）
            headers = [cell.value for cell in sheet[3]]  # 假设第二行是列名

            # 创建一个空的列表，用于存储提取的数据
            extracted_rows = []

            # 遍历工作表中的每一行
            for row in sheet.iter_rows(min_row=4, values_only=True):  # 从第三行开始读取数据
                row_data = dict(zip(headers, row))
                extracted_row = {
                    '序号': row_data.get('序号', ''),
                    '发布媒体': row_data.get('发布媒体', ''),
                    '媒体类型': row_data.get('媒体类型', ''),
                    '稿件名称': row_data.get('稿件名称', ''),
                    '相关链接': row_data.get('相关链接', '')
                }
                extracted_rows.append(extracted_row)

            # 将提取的数据转换为DataFrame
            extracted_data = pd.DataFrame(extracted_rows)

            # 特殊处理“相关链接”列，提取图片并保存
            for index, row in extracted_data.iterrows():
                if '相关链接' in row and row['相关链接']:
                    link = row['相关链接']
                    if isinstance(link, str):
                        # 如果是图片链接，直接保存
                        extracted_data.at[index, '相关链接'] = link
                    else:
                        # 如果是图片对象，提取图片并保存
                        if hasattr(link, 'ref'):
                            image_path = os.path.join(image_folder, f"{file_name}_{index}.png")
                            with open(image_path, 'wb') as f:
                                f.write(link._data())
                            extracted_data.at[index, '相关链接'] = image_path

            # 将提取的数据追加到all_data中
            all_data = pd.concat([all_data, extracted_data], ignore_index=True)

        except Exception as e:
            print(f"Error processing file {file_name}: {e}")

# 将合并后的数据保存到新的Excel文件中
all_data.to_excel(output_file, index=False)

print(f"All data has been merged and saved to {output_file}")
print(f"All images have been saved to {image_folder}")